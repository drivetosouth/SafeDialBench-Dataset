import json
import glob
import re
import os
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

VALID_EVAL_TYPES = {'attack', 'ethics', 'fair', 'legal', 'moral', 'privacy'}
INDUCTION_TYPE_MAP = {
    '场景构建': 'Scenario Construction',
    '话题转换': 'Topic Inversion',
    '角色扮演': 'Role Playing',
    '谬误攻击失败': 'Fallacy Attack',
    '目的反转': 'Purpose Inversion',
    '上下文幻觉': 'Context Hallucination',
    '试探性提问': 'Probing Questions',
    '指代消解': 'Reference Resolution'
}

def extract_model_from_filename(filename):
    base_name = os.path.basename(filename)
    pattern = r"^gpt-3\.5-turbo_([\w.-]+?)(?:[-_](en))?-(\w+)\.jsonl$"
    match = re.match(pattern, base_name)

    if match:
        base_model = match.group(1).strip('-_')
        lang = 'en' if match.group(2) else 'zh'
        eval_type = match.group(3).lower()

        if eval_type not in VALID_EVAL_TYPES:
            print(f"Warning: {base_name} contains invalid evaluation type {eval_type}")
            return (None, 'invalid')
        return (base_model, lang)

    print(f"Filename parsing failed: {base_name}")
    return (None, 'unknown')


def standardize_model_name(raw_name):
    standardized = re.sub(r'[-_]en$', '', raw_name.strip())
    return re.sub(r'_+', '-', standardized)

def is_valid_scores(scores):
    return not all(abs(score - 1.0) < 1e-6 for score in scores.values())

def add_serial_number(df, group_col='Model'):
    if df.empty:
        df.insert(0, 'No.', '')
        return df

    df = df.sort_values(group_col).reset_index(drop=True)
    df = df.copy()

    df['_group_flag'] = (df[group_col] != df[group_col].shift(1)).cumsum()

    serials = []
    current_group = None
    counter = 0

    for idx, row in df.iterrows():
        if row[group_col] != current_group:
            counter += 1
            current_group = row[group_col]
            serials.append(counter)
        else:
            serials.append('')

    df['No.'] = serials
    return df.drop(columns=['_group_flag']).reset_index(drop=True)

# Core Processing Functions
def process_data(file_paths):
    zh_model_data = defaultdict(lambda: {
        'identification': [], 'handling': [], 'consistency': [],
        'count': 0, 'filtered_count': 0,
        'induction_types': defaultdict(int),
        'type_scores': defaultdict(lambda: {'identification': [], 'handling': [], 'consistency': []})
    })
    en_model_data = defaultdict(lambda: zh_model_data.default_factory())
    all_model_data = defaultdict(lambda: zh_model_data.default_factory())

    total_count = 0
    error_count = 0

    for file_path in file_paths:
        raw_model, lang = extract_model_from_filename(file_path)
        if raw_model == "invalid":
            continue

        true_model = standardize_model_name(raw_model)
        print(f"Processing file: {os.path.basename(file_path)} → Model: {true_model} Language: {lang}")

        with open(file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                total_count += 1
                try:
                    item = json.loads(line)
                except json.JSONDecodeError as e:
                    error_count += 1
                    print(f" Line {line_num} JSON error: {str(e)}")
                    continue

                item["model"] = true_model

                required_fields = ["model", "judge", "identification_score",
                                 "handling_score", "consistency_score"]
                if not all(k in item for k in required_fields):
                    error_count += 1
                    print(f" Line {line_num} missing required fields")
                    continue

                try:
                    scores = {
                        'identification': float(item["identification_score"]),
                        'handling': float(item["handling_score"]),
                        'consistency': float(item["consistency_score"])
                    }
                except (ValueError, TypeError) as e:
                    error_count += 1
                    print(f" Line {line_num} score error: {str(e)}")
                    continue

                eval_type = next(
                    (e.replace('-evaluation', '') for e in item["judge"]
                     if e != "gpt-3.5-turbo" and e.replace('-evaluation', '') in VALID_EVAL_TYPES),
                    None
                )
                if not eval_type:
                    print(f" Line {line_num} invalid evaluation type: {item['judge']}")
                    continue

                if not is_valid_scores(scores):
                    targets = [zh_model_data, en_model_data, all_model_data] if lang == 'zh' else [en_model_data, zh_model_data, all_model_data]
                    for data_dict in targets:
                        data_dict[(true_model, eval_type)]['filtered_count'] += 1
                    continue

                key = (true_model, eval_type)
                induction_type = INDUCTION_TYPE_MAP.get(item.get("method", ""), 'Unknown Type')

                # Update corresponding dataset
                target_data = zh_model_data if lang == 'zh' else en_model_data
                for data_dict in [target_data, all_model_data]:
                    data_dict[key]['count'] += 1
                    for metric in scores:
                        data_dict[key][metric].append(scores[metric])
                    data_dict[key]['induction_types'][induction_type] += 1
                    for metric in scores:
                        data_dict[key]['type_scores'][induction_type][metric].append(scores[metric])

    print(f"\nData Processing Summary:")
    print(f"Total processed records: {total_count} | Valid records: {total_count - error_count} | Error data: {error_count}")
    return zh_model_data, en_model_data, all_model_data

def generate_reports(model_data, lang_suffix=''):
    overall_report = []
    for (model, eval_type), data in model_data.items():
        row = {
            'Model': model,
            'Evaluation Type': eval_type,
            'Identification Avg': round(sum(data['identification'])/len(data['identification']), 2) if data['identification'] else 0,
            'Handling Avg': round(sum(data['handling'])/len(data['handling']), 2) if data['handling'] else 0,
            'Consistency Avg': round(sum(data['consistency'])/len(data['consistency']), 2) if data['consistency'] else 0,
            'Valid Samples': data['count'],
            'Filtered Samples': data['filtered_count'],
            'Total Samples': data['count'] + data['filtered_count']
        }
        overall_report.append(row)

    type_report = []
    for (model, eval_type), data in model_data.items():
        total = data['count']
        for itype, count in data['induction_types'].items():
            scores = data['type_scores'][itype]
            row = {
                'Model': model,
                'Evaluation Type': eval_type,
                'Induction Type': itype,
                'Identification Avg': round(sum(scores['identification'])/len(scores['identification']), 2) if scores['identification'] else 0,
                'Handling Avg': round(sum(scores['handling'])/len(scores['handling']), 2) if scores['handling'] else 0,
                'Consistency Avg': round(sum(scores['consistency'])/len(scores['consistency']), 2) if scores['consistency'] else 0,
                'Occurrences': count,
                'Percentage(%)': round(count / total * 100, 1) if total else 0
            }
            type_report.append(row)

    return pd.DataFrame(overall_report), pd.DataFrame(type_report)

def format_excel(input_path, output_path):
    """Excel format optimization"""
    wb = openpyxl.load_workbook(input_path)

    for sheet in wb:
        merge_model_cells(sheet)
        set_global_style(sheet)
        auto_adjust_columns(sheet)

    wb.save(output_path)

def merge_model_cells(sheet):
    current_model = None
    start_row = 2

    for row_idx in range(2, sheet.max_row + 1):
        model = sheet[f'B{row_idx}'].value

        if model == current_model:
            continue

        if current_model is not None:
            end_row = row_idx - 1
            if start_row < end_row:
                sheet.merge_cells(f'A{start_row}:A{end_row}')
                sheet.merge_cells(f'B{start_row}:B{end_row}')

        current_model = model
        start_row = row_idx

    if current_model and start_row < sheet.max_row:
        sheet.merge_cells(f'A{start_row}:A{sheet.max_row}')
        sheet.merge_cells(f'B{start_row}:B{sheet.max_row}')

def set_global_style(sheet):
    font = Font(name='Times New Roman', size=12)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = align

def auto_adjust_columns(sheet):
    for col in sheet.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col),
            default=0
        )
        adjusted_width = (max_len + 2) * 1.2
        sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

def main():
    file_paths = glob.glob("your_data_path/*.jsonl") + \
                 glob.glob("your_data_path/*_en.jsonl")

    if not file_paths:
        print("Error: No input files found")
        return

    print(f"Found {len(file_paths)} data files")

    zh_data, en_data, all_data = process_data(file_paths)

    df_zh_overall, df_zh_type = generate_reports(zh_data, '_cn')
    df_en_overall, df_en_type = generate_reports(en_data, '_en')
    df_all_overall, df_all_type = generate_reports(all_data, '_all')

    overall_cols = ['No.', 'Model', 'Evaluation Type', 'Identification Avg',
                   'Handling Avg', 'Consistency Avg',
                   'Valid Samples', 'Filtered Samples', 'Total Samples']
    type_cols = ['No.', 'Model', 'Evaluation Type', 'Induction Type',
                'Identification Avg', 'Handling Avg', 'Consistency Avg',
                'Occurrences', 'Percentage(%)']

    with pd.ExcelWriter('evaluation_results_raw.xlsx') as writer:
        # Write Chinese data
        df_zh_overall_ordered = add_serial_number(df_zh_overall.sort_values(by=['Model', 'Evaluation Type']))
        df_zh_overall_ordered[overall_cols].to_excel(writer, sheet_name='Overall_Chinese', index=False)
        for itype in list(INDUCTION_TYPE_MAP.values()) + ['Unknown Type']:
            df_sub = df_zh_type[df_zh_type['Induction Type'] == itype]
            if not df_sub.empty:
                df_sub = add_serial_number(df_sub.sort_values(by=['Model', 'Evaluation Type']))
                sheet_name = f"{itype[:28]}_Chinese" if len(itype) > 28 else f"{itype}_Chinese"
                df_sub[type_cols].to_excel(writer, sheet_name=sheet_name[:31], index=False)

        df_en_overall_ordered = add_serial_number(df_en_overall.sort_values(by=['Model', 'Evaluation Type']))
        df_en_overall_ordered[overall_cols].to_excel(writer, sheet_name='Overall_English', index=False)
        for itype in list(INDUCTION_TYPE_MAP.values()) + ['Unknown Type']:
            df_sub = df_en_type[df_en_type['Induction Type'] == itype]
            if not df_sub.empty:
                df_sub = add_serial_number(df_sub.sort_values(by=['Model', 'Evaluation Type']))
                sheet_name = f"{itype[:28]}_English" if len(itype) > 28 else f"{itype}_English"
                df_sub[type_cols].to_excel(writer, sheet_name=sheet_name[:31], index=False)

        df_all_overall_ordered = add_serial_number(df_all_overall.sort_values(by=['Model', 'Evaluation Type']))
        df_all_overall_ordered[overall_cols].to_excel(writer, sheet_name='Overall_Combined', index=False)
        for itype in list(INDUCTION_TYPE_MAP.values()) + ['Unknown Type']:
            df_sub = df_all_type[df_all_type['Induction Type'] == itype]
            if not df_sub.empty:
                df_sub = add_serial_number(df_sub.sort_values(by=['Model', 'Evaluation Type']))
                sheet_name = f"{itype[:28]}_Combined" if len(itype) > 28 else f"{itype}_Combined"
                df_sub[type_cols].to_excel(writer, sheet_name=sheet_name[:31], index=False)

    format_excel('evaluation_results_raw.xlsx', 'evaluation_results_final.xlsx')
    print("\nProcessing complete, final report saved to: evaluation_results_final.xlsx")

if __name__ == "__main__":
    main()